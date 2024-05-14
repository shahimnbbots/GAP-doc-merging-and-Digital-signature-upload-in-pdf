import os
import shutil
import time
import fitz
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import win32com.client as win32
import tkinter as tk


def test(Digi_sign):
    # Create the GUI windows
    root = tk.Tk()
    root.title(Digi_sign)  # Set the title of the window
    status_label = tk.Label(root, text="Started...", padx=30, pady=15)
    status_label.pack()
    input_folder = "D:/DIGITAL SIGN(INV&PL)"
    output_folder = "D:/DIGITAL INV&PL (After Sign)"
    output_folder_1 = "D:/GAP_MERGING_FILES(BOT)"
    sign_file = "Signature.png"

    # Create the output folders if they don't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    if not os.path.exists(output_folder_1):
        os.makedirs(output_folder_1)

    # Iterate through each PDF file in the input folder
    for filename in os.listdir(input_folder):
        if filename.endswith(".pdf"):
            input_file = os.path.join(input_folder, filename)
            output_file = os.path.join(output_folder, filename)
            output_file_1 = os.path.join(output_folder_1, filename)
            status_label.config(text=f"Processing pdf: {filename}")
            status_label.update()  # Update the GUI to reflect the changes
            time.sleep(1)
            # Open the PDF file
            file_handle = fitz.open(input_file)
            # Text to search for
            search_text = "I hereby certify that all information provided is true and correct."
            status_label.config(text=f"Text Found: {search_text}")
            status_label.update()  # Update the GUI to reflect the changes
            time.sleep(1)
            # Iterate through each page and search for the text
            for page_num in range(file_handle.page_count):
                page = file_handle.load_page(page_num)
                text_instances = page.search_for(search_text)
                if text_instances:
                    text_rectangle = text_instances[0]

                    # Calculate the position to insert the image right after the text
                    x = text_rectangle[2] + 5
                    y = text_rectangle[1] + 20  # Maintain the same y-coordinate as the text

                    # Define the rectangle for the image
                    image_rectangle = fitz.Rect(x, y, x + 130, y + 100)  # Adjust size as needed
                    # Insert the image
                    page.insert_image(image_rectangle, filename=sign_file)
                    # Save the modified PDF to the first output location
                    file_handle.save(output_file)
                    # Save the modified PDF to the second output location
                    shutil.copy(output_file, output_file_1)
                    status_label.config(text=f"Image inserted successfully in {filename}.")
                    status_label.update()  # Update the GUI to reflect the changes
                    time.sleep(1)
                    print(f"Image inserted successfully in {filename}.")
                    break  # Stop searching after finding the text once
            # Close the PDF file
            file_handle.close()
            status_label.config(text=f"Modified PDF saved to {output_file} and {output_file_1}.")
            status_label.update()  # Update the GUI to reflect the changes
            time.sleep(1)
            print(f"Modified PDF saved to {output_file} and {output_file_1}.")
    status_label.config(text=f"All PDF files processed.")
    status_label.update()  # Update the GUI to reflect the changes
    time.sleep(1)
    print("All PDF files processed.")

    # Initialize Excel application
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False  # Set to True if you want Excel to be visible during execution

    print("Processing Excel files...")
    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            status_label.config(text=f"Found excel {filename}")
            status_label.update()  # Update the GUI to reflect the changes
            time.sleep(1)
            print(f"Found Excel file: {filename}")
            # Open the Excel file
            excel_file = os.path.join(input_folder, filename)
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Find the row with the text "Preparer’s Company Chop/Stamp" in column A
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    if "Preparer’s Company Chop/Stamp" in str(cell.value):
                        status_label.config(text=f"Found text: Preparer’s Company Chop/Stamp")
                        status_label.update()  # Update the GUI to reflect the changes
                        time.sleep(1)
                        print("Found text: Preparer’s Company Chop/Stamp")

                        # Insert the image after the text
                        image_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        img = ExcelImage(sign_file)
                        # Resize the image (adjust as needed)
                        img.width = 200
                        img.height = 60
                        img.anchor = f"{image_cell.column_letter}{image_cell.row - 1}"
                        sheet.add_image(img)

                        # Find the PO number in cell B12
                        po_number = sheet['B12'].value
                        if po_number:
                            status_label.config(text=f"Found po number:{po_number}")
                            status_label.update()  # Update the GUI to reflect the changes
                            time.sleep(1)
                            print(f"Found PO number: {po_number}")
                            # Save the modified Excel file with the PO number as the filename in the output folder
                            output_excel_file = os.path.join(output_folder, f"{po_number}-SUMMARY.xlsx")
                            wb.save(output_excel_file)
                            # wb.close()  # Close the workbook without saving changes
                            # Convert the Excel file to PDF
                            pdf_output_file = os.path.join(output_folder, f"{po_number}-SUMMARY.pdf")
                            excel.Workbooks.Open(output_excel_file).ExportAsFixedFormat(0, pdf_output_file)
                            # Close the workbook
                            excel.ActiveWorkbook.Close(SaveChanges=True)
                            wb.close()
                            # Wait for a few seconds to ensure the file is closed
                            time.sleep(3)  # Remove the Excel file
                            status_label.config(text=f"Renamed{pdf_output_file}")
                            status_label.update()  # Update the GUI to reflect the changes
                            time.sleep(1)
                            os.remove(output_excel_file)
                            status_label.config(text=f"Image inserted, Excel file converted to PDF, and Excel file removed: {output_excel_file}")
                            status_label.update()  # Update the GUI to reflect the changes
                            print(f"Image inserted, Excel file converted to PDF, and Excel file removed: {output_excel_file}")
                            shutil.copy(pdf_output_file, output_folder_1)
                        else:
                            status_label.config(text=f"PO number not found in cell B12")
                            status_label.update()  # Update the GUI to reflect the changes
                            time.sleep(1)
                            print("PO number not found in cell B12")
                        break  # Stop searching after finding the text once
    status_label.config(text=f"All Excel files processed.")
    status_label.update()  # Update the GUI to reflect the changes
    time.sleep(1)
    status_label.config(text=f"Completed")
    status_label.update()  # Update the GUI to reflect the changes
    time.sleep(1)
    print("All Excel files processed.")
    print("Completed")
    root.destroy()


test("Digi_sign")